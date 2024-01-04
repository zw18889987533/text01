"""
编写函数，作用为excle数据处理

"""
import math

import xlrd as rd
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl import load_workbook


def open_excle_data(file, sheet, key, hang, lie):
    """
    读取此sheet页，所有的数据，并输出成字典
    :param file: 文件路径
    :param sheet: 要抓取的sheet页
    :param key: key的行数（表头）
    :param hang: 从哪行开始抓取
    :param lie: 从哪列开始抓取
    :return: 返回一个字典｛id:{'备注'：xxx,'初始星级‘：xxx,...}｝
    """
    # 打开表
    book = rd.open_workbook(file)
    # 找到sheet
    sheet_1 = book.sheet_by_index(sheet)  # 获取不同sheet页

    # 获取总行总列
    row_Num = sheet_1.nrows
    col_Num = sheet_1.ncols

    data_list = []

    # 取Key元素
    key_list = sheet_1.row_values(key)

    # 从13行开始读取数据
    j = hang
    for i in range(hang, row_Num):
        dict = {}
        # 每一行数据，取出来作为values
        values = sheet_1.row_values(j)
        for x in range(lie, col_Num):
            # 拿key列表循环与values循环插入{key_list[x]:values[x]}
            dict[key_list[x]] = values[x]
        j += 1

        # 把字典加入到列表中
        data_list.append(dict)

    # 输出表的第一列ID
    id_list = []
    for i in range(hang, row_Num):
        id = int(sheet_1.cell_value(i, 0))
        id_list.append(id)
    # print(id_list)

    # 输出整个文档sheet的字典数据
    dict_result = {}
    for i in range(len(id_list)):
        dict_result[id_list[i]] = data_list[i]

    return dict_result


def open_excle_DB(file1, file2, sheet, file3):
    """

    :param file1:最新的文件
    :param file2:上一个版本的文件
    :param sheet:哪个sheet
    :param file3:输出路径
    :return:返回输出的文件
    """
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    sheet1 = wb1[sheet]
    sheet2 = wb2[sheet]
    # 遍历两个表中的最大行，取更大的那一个
    max_row = sheet1.max_row if sheet1.max_row > sheet2.max_row else sheet2.max_row

    # 遍历所有的列，取更大的那一个
    max_column = sheet1.max_column if sheet1.max_column > sheet2.max_column else sheet2.max_column

    # 使用for循环分别遍历行数据与列数据，判断对于单元格数值是否相等
    for i in range(1, (max_row + 1)):
        for j in range(1, (max_column + 1)):
            cell_1 = sheet1.cell(i, j)
            cell_2 = sheet2.cell(i, j)
            if cell_1.value != cell_2.value:
                # cell_1.fill = PatternFill("solid", fgColor='FFFF00')
                # cell_1.font = Font(color=colors.BLACK, bold=True)
                cell_2.fill = PatternFill("solid", fgColor='FFFF00')
                cell_2.font = Font(color=colors.BLACK, bold=True)

    wb2.save(file3)


def open_item():
    """
    #
    :return:{道具类型：[id,name,drop]}
    """
    book = rd.open_workbook('/文件处理/配置数据表/Item.xlsx')
    sheet_0 = book.sheet_by_index(0)
    row_Num = sheet_0.nrows
    col_Num = sheet_0.ncols
    # 把物品类型列表拿出来作为KYE
    key_list = []
    sheet_4 = book.sheet_by_index(4)
    for i in range(5, 23):
        data = sheet_4.cell_value(i, 1)
        key_list.append(data)

    dict = {}
    for i in range(13, row_Num):
        id = int(sheet_0.cell_value(i, 0))
        name = str(sheet_0.cell_value(i, 2))
        drop = str(sheet_0.cell_value(i, 6))
        ItemType = sheet_0.cell_value(i, 5)
        # item_name_list.append(name)
        # item_id_list.append(id)
        # item_DropGroup.append(drop)

        try:
            dict[ItemType].append([id, name, drop])

        except KeyError:
            dict[ItemType] = []
            dict[ItemType].append([id, name, drop])

    return dict


def pandas_open_file(file, header, skiprows, usecols, sheet):
    """
    使用pandas读取文件输出至另外一个文档
    :param file: 读取文件
    :param header: 读取文件头
    :param skiprows: 从多少行开始读取
    :param usecols: 多少列到多少列
    :param sheet: 抓取的sheet页
    :return:
    """
    head = pd.read_excel(file, sheet_name=sheet, header=header, usecols=usecols, index_col=None)
    head_list = head.columns
    people = pd.read_excel(file, sheet_name=sheet, skiprows=skiprows, usecols=usecols, index_col=None)
    people.columns = head_list
    # people.to_excel(excel_file)
    return people


def dict_key_value(data_dict):
    # 遍历字典中的每个键值对
    """
    把数据｛id:xxx:'buf'｝去除key和对应的所有value，组合成{61201750: [742017500, 742017502, 742017507], 61201645: [742016412, 742016413, 700004040]}
    :param data_dict:
    :return:
    """
    dict1 = {}
    dict2 = {}
    dict3 = {}
    dict4 = {}
    for key, value in data_dict.items():
        # 提取字典内的值，并生成一个列表，去除空值
        values_list = [str(val) for val in value.values() if val != '']
        # 将键和列表生成一个新的字典
        dict1[key] = values_list

    # 数据处理，把/和.0去除
    for key, values in dict1.items():
        new_values = []
        for value in values:
            if '/' in value:
                new_values.extend(value.split('/'))
            else:
                new_values.append(value)
        new_values = [v.replace('.0', '') for v in new_values]
        dict2[key] = new_values
    # value值取整数
    for key, values in dict2.items():
        new_values1 = [int(value) for value in values]
        dict3[key] = new_values1

    # 删除相同的buf
    for key, values in dict3.items():
        unique_values = list(set(values))
        if unique_values:
            dict4[key] = unique_values
    return dict4


def buf_data(buf_dict, book, data_list):
    for sillk_key, sillk_values in buf_dict.items():
        cs_buf = sillk_values  # list
        cs_buf_data = {sillk_key: book[sillk_key] for sillk_key in cs_buf if sillk_key in book}
        cs_buf_sx = {}
        for buf_key, buf_value in cs_buf_data.items():
            cs_buf_sx[buf_key] = {
                'Desc': buf_value['Desc'],
                'LifeTime': buf_value['LifeTime'],
                'FuncType': buf_value['FuncType'],
                'FuncParams': buf_value['FuncParams'],
                'ActiveCondIds': buf_value['ActiveCondIds'],
                'PropertiesList': buf_value['PropertiesList'],
            }

            buf_pack = cs_buf_sx[buf_key]['FuncType']
            if buf_pack == 'Pack':
                buf_Func_list = []
                buf_Func = cs_buf_sx[buf_key]['FuncParams']
                buf_Func_list.append(buf_Func)
                # print('buff_Pack:', buf_key)
                # print(buf_Func_list)
                # 包内的buf列表
                lst = [int(x) for item in buf_Func_list for x in item.split('/')]
                pack_buf_data = {buf_key: book[buf_key] for buf_key in lst if buf_key in book}
                # print('pack内buf数据：', pack_buf_data)
                pack_buf_sx = {}
                for k, v in pack_buf_data.items():
                    pack_buf_sx[k] = {
                        'Desc': v['Desc'],
                        'LifeTime': v['LifeTime'],
                        'FuncType': v['FuncType'],
                        'FuncParams': v['FuncParams'],
                        'ActiveCondIds': v['ActiveCondIds'],
                        'PropertiesList': v['PropertiesList'],
                    }

                # print('pack内buf:',pack_buf_sx)
                buf_sx_dict = {}
                for key, value in pack_buf_sx.items():
                    buf_sx_dict[key] = value
                # print(buf_sx_dict)
                for pack_buf_key, pack_buf_value in buf_sx_dict.items():
                    pack_buf_dict = []
                    pack_buf_dict.append(sillk_key)  # 技能的id
                    pack_buf_dict.append(pack_buf_key)  # buf的包id
                    pack_buf_dict.append(buf_key)  # buf的id
                    pack_buf_dict.append(pack_buf_value)  # buf里的参数
                    data_list.append(pack_buf_dict)
                    # print(pack_buf_dict)
            else:
                pass

        # buf_data_dict['buf_pack'] = 0
        # buf_num = len(list(cs_buf_sx.values()))

        for key_id in cs_buf_sx.keys():
            buf_data_list = []
            buf_data_list.append(sillk_key)
            # print('id',key_id)
            buf_data_list.append(key_id)
            buf_data_list.append('')
            buf_data_list.append(cs_buf_sx[key_id].values())
            dict_values_list = list(buf_data_list[-1])
            buf_data_list[-1] = {
                'Desc': dict_values_list[0],
                'LifeTime': dict_values_list[1],
                'FuncType': dict_values_list[2],
                'FuncParams': dict_values_list[3],
                'ActiveCondIds': dict_values_list[4],
                'PropertiesList': dict_values_list[5],
            }
            data_list.append(buf_data_list)
    return data_list


def replace_nan(val, ActorBuf_data):
    """
    处理数据
    :param val:
    :param ActorBuf_data:
    :return:
    """
    if isinstance(val, list):
        return [replace_nan(sub_val) for sub_val in val]
    elif isinstance(val, str):
        try:
            val = float(val)
        except ValueError:
            return val
    if isinstance(val, float) and math.isnan(val):
        return ''
    else:
        return ActorBuf_data.get(int(val), {}).get('Desc', '')


if __name__ == '__main__':
    # a = open_excle_data('D:\pythonProject\文件处理\配置数据表/英雄基础信息表.xlsx', 0, 4, 12, 1)
    # # print(a)
    # name_list = []
    # for key in a:
    #     name = a[key]['碎片道具ID']
    #     name_list.append(name)
    # # print(name_list)
    pass
